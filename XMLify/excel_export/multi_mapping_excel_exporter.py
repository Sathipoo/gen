import pandas as pd
import networkx as nx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from typing import Dict, List, Any, Optional
from collections import defaultdict, Counter
import json
import os

class MultiMappingExcelExporter:
    """
    Enhanced Excel exporter for multiple Informatica mappings in a single Excel file
    """
    
    def __init__(self, output_path: str = None):
        """
        Initialize with optional output path
        
        Args:
            output_path: Path for output Excel file (optional for first initialization)
        """
        self.output_path = output_path
        self.workbook = None
        self.is_new_file = True
        
        # Define color schemes
        self.colors = {
            'header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'source': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            'target': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),
            'transformation': PatternFill(start_color='F0E68C', end_color='F0E68C', fill_type='solid'),
            'parallel': PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid'),
            'critical': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
            'light_gray': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
            'mapping_header': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
            'separator': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        }
        
        self.font_bold = Font(bold=True, color='FFFFFF')
        self.font_regular = Font(color='000000')
        self.font_mapping_header = Font(bold=True, color='FFFFFF', size=12)
        
        # Track current row positions for each sheet
        self.sheet_positions = {
            'overview': 1,
            'detailed_lineage': 1,
            'transformation_logic': 1,
            'execution_plan': 1,
            'data_flow': 1,
            'performance_analysis': 1
        }
        
        # Initialize workbook
        self._initialize_workbook()
    
    def _initialize_workbook(self):
        """Initialize or load existing workbook"""
        if self.output_path and os.path.exists(self.output_path):
            # Load existing workbook
            self.workbook = load_workbook(self.output_path)
            self.is_new_file = False
            print(f"📁 Loaded existing Excel file: {self.output_path}")
        else:
            # Create new workbook
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)  # Remove default sheet
            self.is_new_file = True
            print("📁 Created new Excel workbook")
    
    def add_mapping_to_excel(self, lineage_generator, output_path: str = None):
        """
        Add a single mapping to the Excel file
        
        Args:
            lineage_generator: InfaLineageGenerator instance
            output_path: Path for output Excel file (optional, uses instance path if not provided)
        """
        if output_path:
            self.output_path = output_path
        
        if self.is_new_file:
            # First mapping - create all sheets
            self._create_overview_sheet()
            self._create_detailed_lineage_sheet()
            self._create_transformation_logic_sheet()
            self._create_execution_plan_sheet()
            self._create_data_flow_sheet()
            self._create_performance_analysis_sheet()
            self.is_new_file = False
        
        # Add mapping data to each sheet
        self._add_mapping_to_overview(lineage_generator)
        self._add_mapping_to_detailed_lineage(lineage_generator)
        self._add_mapping_to_transformation_logic(lineage_generator)
        self._add_mapping_to_execution_plan(lineage_generator)
        self._add_mapping_to_data_flow(lineage_generator)
        self._add_mapping_to_performance_analysis(lineage_generator)
        
        # Save workbook
        self.workbook.save(self.output_path)
        print(f"✅ Added mapping '{lineage_generator.mapping_name}' to Excel file")
    
    def _get_or_create_sheet(self, sheet_name: str):
        """Get existing sheet or create new one"""
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return self.workbook.create_sheet(sheet_name)
    
    def _create_overview_sheet(self):
        """Create overview sheet with header"""
        ws = self._get_or_create_sheet("📊 Overview")
        
        # Title
        ws['A1'] = "Informatica Multi-Mapping Lineage Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = "Comprehensive Analysis of All Mappings"
        ws['A2'].font = Font(size=14, bold=True)
        
        # Headers for mapping data
        headers = [
            "Mapping Name", "Total Transformations", "Total Connections", 
            "Source Definitions", "Target Definitions", "Expression Transformations",
            "Filter Transformations", "Aggregator Transformations", "Lookup Transformations",
            "Components", "Max Execution Level", "Parallel Efficiency %"
        ]
        
        for j, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=j, value=header)
            cell.fill = self.colors['header']
            cell.font = self.font_bold
        
        self.sheet_positions['overview'] = 5  # Start data from row 5
    
    def _add_mapping_to_overview(self, lg):
        """Add mapping data to overview sheet"""
        ws = self.workbook["📊 Overview"]
        row = self.sheet_positions['overview']
        
        # Add mapping separator
        if row > 5:  # Not the first mapping
            for col in range(1, 13):  # 12 columns
                cell = ws.cell(row=row, column=col)
                cell.fill = self.colors['separator']
            row += 1
        
        # Add mapping header
        ws[f'A{row}'] = f"📋 {lg.mapping_name}"
        ws[f'A{row}'].font = self.font_mapping_header
        ws[f'A{row}'].fill = self.colors['mapping_header']
        row += 1
        
        # Calculate statistics
        G = lg.mapping_lineage_graph
        connected_components = list(nx.weakly_connected_components(G))
        
        # Count transformation types
        type_counts = defaultdict(int)
        for instance in lg.instances:
            trans_type = instance.get('@TRANSFORMATION_TYPE', 'Unknown')
            type_counts[trans_type] += 1
        
        # Calculate parallel efficiency
        max_level = max(lg.transformation_orders.values()) if lg.transformation_orders else 0
        total_transformations = len([t for t in lg.transformation_orders.values() if t > 0])
        parallel_efficiency = (total_transformations / max_level * 100) if max_level > 0 else 0
        
        # Add data row
        data = [
            lg.mapping_name,
            len(lg.instances),
            len(lg.connectors),
            type_counts.get('Source Definition', 0),
            type_counts.get('Target Definition', 0),
            type_counts.get('Expression', 0),
            type_counts.get('Filter', 0),
            type_counts.get('Aggregator', 0),
            type_counts.get('Lookup Procedure', 0),
            len(connected_components),
            max_level,
            f"{parallel_efficiency:.1f}"
        ]
        
        for j, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=j, value=value)
            if row % 2 == 0:
                cell.fill = self.colors['light_gray']
        
        self.sheet_positions['overview'] = row + 2  # Update position for next mapping
    
    def _create_detailed_lineage_sheet(self):
        """Create detailed lineage sheet with header"""
        ws = self._get_or_create_sheet("🔗 Detailed Lineage")
        
        # Title
        ws['A1'] = "Detailed Lineage Analysis - All Mappings"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Headers
        headers = [
            "Mapping Name", "From Field", "From Instance", "From Type", 
            "To Field", "To Instance", "To Type", "Execution Order", 
            "Execution Level", "Parallel Group", "Transformation Logic"
        ]
        
        for j, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=j, value=header)
            cell.fill = self.colors['header']
            cell.font = self.font_bold
        
        self.sheet_positions['detailed_lineage'] = 4  # Start data from row 4
    
    def _add_mapping_to_detailed_lineage(self, lg):
        """Add mapping lineage data to detailed lineage sheet"""
        ws = self.workbook["🔗 Detailed Lineage"]
        row = self.sheet_positions['detailed_lineage']
        
        # Add mapping header
        ws[f'A{row}'] = f"📋 {lg.mapping_name}"
        ws[f'A{row}'].font = self.font_mapping_header
        ws[f'A{row}'].fill = self.colors['mapping_header']
        row += 1
        
        # Add connector data
        for connector in lg.connectors_data:
            data = [
                lg.mapping_name,
                connector.from_field,
                connector.from_instance,
                connector.from_instance_type,
                connector.to_field,
                connector.to_instance,
                connector.to_instance_type,
                connector.transformation_order,
                connector.execution_level,
                connector.parallel_group,
                connector.transformation_logic[:100] + "..." if len(connector.transformation_logic) > 100 else connector.transformation_logic
            ]
            
            for j, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=j, value=value)
                
                # Color code by transformation type
                if 'Source' in str(data[3]):
                    cell.fill = self.colors['source']
                elif 'Target' in str(data[6]):
                    cell.fill = self.colors['target']
                elif row % 2 == 0:
                    cell.fill = self.colors['light_gray']
            
            row += 1
        
        self.sheet_positions['detailed_lineage'] = row + 1  # Add space between mappings
    
    def _create_transformation_logic_sheet(self):
        """Create transformation logic sheet with header"""
        ws = self._get_or_create_sheet("⚙️ Transformation Logic")
        
        # Title
        ws['A1'] = "Transformation Logic - All Mappings"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Headers
        headers = ["Mapping Name", "Transformation Name", "Type", "Execution Order", "Business Logic", "Field Count"]
        for j, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=j, value=header)
            cell.fill = self.colors['header']
            cell.font = self.font_bold
        
        self.sheet_positions['transformation_logic'] = 4
    
    def _add_mapping_to_transformation_logic(self, lg):
        """Add transformation logic data"""
        ws = self.workbook["⚙️ Transformation Logic"]
        row = self.sheet_positions['transformation_logic']
        
        # Add mapping header
        ws[f'A{row}'] = f"📋 {lg.mapping_name}"
        ws[f'A{row}'].font = self.font_mapping_header
        ws[f'A{row}'].fill = self.colors['mapping_header']
        row += 1
        
        if lg.mapping_name in lg.transformation_logic_cache:
            for trans_name, trans_info in lg.transformation_logic_cache[lg.mapping_name].items():
                # Get execution order
                full_trans_name = None
                for instance in lg.instances:
                    if instance.get('@NAME') == trans_name:
                        trans_type = instance.get('@TRANSFORMATION_TYPE', 'Unknown')
                        full_trans_name = f"{lg.create_transform_type_acronym(trans_type)}{trans_name}"
                        break
                
                exec_order = lg.transformation_orders.get(full_trans_name, 0)
                
                # Count fields in transformation
                field_count = 0
                for transformation in lg.transformations:
                    if transformation.get('@NAME') == trans_name:
                        transform_fields = transformation.get('TRANSFORMFIELD', [])
                        if isinstance(transform_fields, dict):
                            field_count = 1
                        elif isinstance(transform_fields, list):
                            field_count = len(transform_fields)
                        break
                
                data = [
                    lg.mapping_name,
                    trans_name,
                    trans_info['type'],
                    exec_order,
                    trans_info['logic'][:200] + "..." if len(trans_info['logic']) > 200 else trans_info['logic'],
                    field_count
                ]
                
                for j, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=j, value=value)
                    if row % 2 == 0:
                        cell.fill = self.colors['light_gray']
                
                row += 1
        
        self.sheet_positions['transformation_logic'] = row + 1
    
    def _create_execution_plan_sheet(self):
        """Create execution plan sheet with header"""
        ws = self._get_or_create_sheet("📋 Execution Plan")
        
        # Title
        ws['A1'] = "Execution Plan - All Mappings"
        ws['A1'].font = Font(size=16, bold=True)
        
        self.sheet_positions['execution_plan'] = 3
    
    def _add_mapping_to_execution_plan(self, lg):
        """Add execution plan data"""
        ws = self.workbook["📋 Execution Plan"]
        row = self.sheet_positions['execution_plan']
        
        # Add mapping header
        ws[f'A{row}'] = f"📋 {lg.mapping_name}"
        ws[f'A{row}'].font = self.font_mapping_header
        ws[f'A{row}'].fill = self.colors['mapping_header']
        row += 1
        
        # Add execution levels
        ws[f'A{row}'] = "Execution Levels:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Group by execution level
        level_groups = defaultdict(list)
        for trans_name, level in lg.transformation_orders.items():
            level_groups[level].append(trans_name)
        
        for level in sorted(level_groups.keys()):
            transformations = level_groups[level]
            if len(transformations) == 1:
                clean_name = transformations[0].split('_', 1)[-1] if '_' in transformations[0] else transformations[0]
                ws[f'A{row}'] = f"   Level {level}: [{clean_name}]"
            else:
                ws[f'A{row}'] = f"   PARALLEL EXECUTION Level {level} ({len(transformations)} transformations):"
                row += 1
                
                for trans_name in transformations:
                    clean_name = trans_name.split('_', 1)[-1] if '_' in trans_name else trans_name
                    ws[f'A{row}'] = f"     • [{clean_name}]"
                    ws[f'A{row}'].fill = self.colors['parallel']
                    row += 1
                row -= 1  # Adjust for the extra increment
            
            row += 1
        
        self.sheet_positions['execution_plan'] = row + 2
    
    def _create_data_flow_sheet(self):
        """Create data flow sheet with header"""
        ws = self._get_or_create_sheet("🔄 Data Flow")
        
        # Title
        ws['A1'] = "Data Flow Visualization - All Mappings"
        ws['A1'].font = Font(size=16, bold=True)
        
        self.sheet_positions['data_flow'] = 3
    
    def _add_mapping_to_data_flow(self, lg):
        """Add data flow information"""
        ws = self.workbook["🔄 Data Flow"]
        row = self.sheet_positions['data_flow']
        
        # Add mapping header
        ws[f'A{row}'] = f"📋 {lg.mapping_name}"
        ws[f'A{row}'].font = self.font_mapping_header
        ws[f'A{row}'].fill = self.colors['mapping_header']
        row += 1
        
        # Add component analysis
        G = lg.mapping_lineage_graph
        connected_components = list(nx.weakly_connected_components(G))
        
        ws[f'A{row}'] = f"Components: {len(connected_components)}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for i, component in enumerate(connected_components):
            sources = len([n for n in component if G.in_degree(n) == 0])
            sinks = len([n for n in component if G.out_degree(n) == 0])
            
            if len(component) == 1:
                comp_type = "Isolated"
                desc = "Standalone transformation (likely lookup)"
            elif sources > 1 and sinks == 1:
                comp_type = "Multi-Source Pipeline"
                desc = f"{sources} sources converging to {sinks} target"
            elif sources == 1 and sinks > 1:
                comp_type = "Multi-Target Pipeline"
                desc = f"{sources} source feeding {sinks} targets"
            else:
                comp_type = "Standard Pipeline"
                desc = f"{sources} source(s) to {sinks} target(s)"
            
            ws[f'A{row}'] = f"   Component {i}: {comp_type} - {desc}"
            row += 1
        
        self.sheet_positions['data_flow'] = row + 2
    
    def _create_performance_analysis_sheet(self):
        """Create performance analysis sheet with header"""
        ws = self._get_or_create_sheet("📉 Performance Analysis")
        
        # Title
        ws['A1'] = "Performance Analysis - All Mappings"
        ws['A1'].font = Font(size=16, bold=True)
        
        self.sheet_positions['performance_analysis'] = 3
    
    def _add_mapping_to_performance_analysis(self, lg):
        """Add performance analysis data"""
        ws = self.workbook["📉 Performance Analysis"]
        row = self.sheet_positions['performance_analysis']
        
        # Add mapping header
        ws[f'A{row}'] = f"📋 {lg.mapping_name}"
        ws[f'A{row}'].font = self.font_mapping_header
        ws[f'A{row}'].fill = self.colors['mapping_header']
        row += 1
        
        # Analyze bottlenecks
        ws[f'A{row}'] = "🚨 Potential Bottlenecks:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        G = lg.mapping_lineage_graph
        bottlenecks = []
        
        # Find nodes with high fan-in (convergence points)
        for node in G.nodes():
            in_degree = G.in_degree(node)
            out_degree = G.out_degree(node)
            
            if in_degree > 2:  # Convergence point
                clean_name = node.split('_', 1)[-1] if '_' in node else node
                bottlenecks.append(f"   High fan-in: {clean_name} ({in_degree} inputs)")
            
            if out_degree > 3:  # High fan-out
                clean_name = node.split('_', 1)[-1] if '_' in node else node
                bottlenecks.append(f"   High fan-out: {clean_name} ({out_degree} outputs)")
        
        for bottleneck in bottlenecks:
            ws[f'A{row}'] = bottleneck
            ws[f'A{row}'].fill = self.colors['critical']
            row += 1
        
        if not bottlenecks:
            ws[f'A{row}'] = "   No significant bottlenecks detected"
            row += 1
        
        # Parallel efficiency
        row += 1
        ws[f'A{row}'] = "⚡ Parallel Efficiency:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        max_level = max(lg.transformation_orders.values()) if lg.transformation_orders else 0
        total_transformations = len([t for t in lg.transformation_orders.values() if t > 0])
        
        if max_level > 0:
            parallel_efficiency = (total_transformations / max_level) * 100
            efficiency_text = f"   Efficiency: {parallel_efficiency:.1f}%"
            
            if parallel_efficiency > 75:
                efficiency_text += " ✅ Good parallelization"
                ws[f'A{row}'] = efficiency_text
                ws[f'A{row}'].fill = self.colors['source']
            elif parallel_efficiency > 50:
                efficiency_text += " ⚠️ Moderate parallelization"
                ws[f'A{row}'] = efficiency_text
                ws[f'A{row}'].fill = self.colors['transformation']
            else:
                efficiency_text += " 🔴 Low parallelization"
                ws[f'A{row}'] = efficiency_text
                ws[f'A{row}'].fill = self.colors['critical']
        else:
            ws[f'A{row}'] = "   Unable to calculate parallel efficiency"
        
        self.sheet_positions['performance_analysis'] = row + 3
    
    def finalize_excel(self):
        """Finalize Excel file with formatting and save"""
        # Auto-adjust column widths for all sheets
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            for column_cells in ws.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        # Save workbook
        self.workbook.save(self.output_path)
        print(f"\n📊 Multi-Mapping Excel Report Finalized!")
        print(f"📁 Location: {self.output_path}")
        print(f"📋 Contains comprehensive analysis of all mappings in 6 sheets")


# Utility functions for easy usage
def create_multi_mapping_exporter(output_path: str):
    """
    Create a new multi-mapping Excel exporter
    
    Args:
        output_path: Path for the Excel file
        
    Returns:
        MultiMappingExcelExporter instance
    """
    return MultiMappingExcelExporter(output_path)

def add_mapping_to_excel(exporter, lineage_generator):
    """
    Add a mapping to the existing Excel exporter
    
    Args:
        exporter: MultiMappingExcelExporter instance
        lineage_generator: InfaLineageGenerator instance
    """
    exporter.add_mapping_to_excel(lineage_generator)

def finalize_multi_mapping_excel(exporter):
    """
    Finalize and save the multi-mapping Excel file
    
    Args:
        exporter: MultiMappingExcelExporter instance
    """
    exporter.finalize_excel() 