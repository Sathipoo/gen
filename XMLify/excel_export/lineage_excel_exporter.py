import pandas as pd
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from typing import Dict, List, Any, Optional
from collections import defaultdict, Counter
import json

class InfaLineageExcelExporter:
    """
    Comprehensive Excel exporter for Informatica mapping lineage analysis
    """
    
    def __init__(self, lineage_generator):
        """
        Initialize with InfaLineageGenerator object
        
        Args:
            lineage_generator: InfaLineageGenerator instance with processed mapping data
        """
        self.lg = lineage_generator
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Define color schemes
        self.colors = {
            'header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'source': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            'target': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),
            'transformation': PatternFill(start_color='F0E68C', end_color='F0E68C', fill_type='solid'),
            'parallel': PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid'),
            'critical': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
            'light_gray': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        }
        
        self.font_bold = Font(bold=True, color='FFFFFF')
        self.font_regular = Font(color='000000')
        
    def generate_excel_report(self, output_path: str):
        """
        Generate comprehensive Excel report with multiple sheets
        """
        print("Generating Excel lineage report...")
        
        # Create all sheets
        self._create_overview_sheet()
        self._create_detailed_lineage_sheet()
        self._create_transformation_logic_sheet()
        self._create_execution_plan_sheet()
        self._create_data_flow_sheet()
        self._create_performance_analysis_sheet()
        
        # Save workbook
        self.workbook.save(output_path)
        print(f"Excel report saved: {output_path}")
        
    def _create_overview_sheet(self):
        """Create overview sheet with mapping summary"""
        ws = self.workbook.create_sheet("📊 Overview")
        
        # Title
        ws['A1'] = f"Informatica Mapping Lineage Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Mapping: {self.lg.mapping_name}"
        ws['A2'].font = Font(size=14, bold=True)
        
        # Basic Statistics
        ws['A4'] = "📈 Mapping Statistics"
        ws['A4'].font = Font(size=12, bold=True)
        
        stats_data = [
            ["Total Transformations", len(self.lg.instances)],
            ["Total Connections", len(self.lg.connectors)],
            ["Source Definitions", len([t for t in self.lg.instances if t.get('@TRANSFORMATION_TYPE') == 'Source Definition'])],
            ["Target Definitions", len([t for t in self.lg.instances if t.get('@TRANSFORMATION_TYPE') == 'Target Definition'])],
            ["Expression Transformations", len([t for t in self.lg.instances if t.get('@TRANSFORMATION_TYPE') == 'Expression'])],
            ["Filter Transformations", len([t for t in self.lg.instances if t.get('@TRANSFORMATION_TYPE') == 'Filter'])],
            ["Aggregator Transformations", len([t for t in self.lg.instances if t.get('@TRANSFORMATION_TYPE') == 'Aggregator'])],
            ["Lookup Transformations", len([t for t in self.lg.instances if t.get('@TRANSFORMATION_TYPE') == 'Lookup Procedure'])],
        ]
        
        for i, (metric, value) in enumerate(stats_data, 5):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            if i % 2 == 0:
                ws[f'A{i}'].fill = self.colors['light_gray']
                ws[f'B{i}'].fill = self.colors['light_gray']
        
        # Component Analysis
        ws['A15'] = "🔄 Component Analysis"
        ws['A15'].font = Font(size=12, bold=True)
        
        # Get connected components info
        G = self.lg.mapping_lineage_graph
        connected_components = list(nx.weakly_connected_components(G))
        
        component_data = [["Component", "Node Count", "Type", "Description"]]
        for i, component in enumerate(connected_components):
            sources = len([n for n in component if G.in_degree(n) == 0])
            sinks = len([n for n in component if G.out_degree(n) == 0])
            
            if len(component) == 1:
                comp_type = "Isolated"
                desc = "Standalone transformation (likely lookup)"
            elif sources > 1 and sinks == 1:
                comp_type = "Multi-Source Pipeline"
                desc = f"{sources} sources converging to {sinks} target"
            elif sources == 1 and sinks > 1:
                comp_type = "Multi-Target Pipeline"
                desc = f"{sources} source feeding {sinks} targets"
            else:
                comp_type = "Standard Pipeline"
                desc = f"{sources} source(s) to {sinks} target(s)"
            
            component_data.append([f"Component {i}", len(component), comp_type, desc])
        
        for i, row in enumerate(component_data, 16):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 16:  # Header
                    cell.fill = self.colors['header']
                    cell.font = self.font_bold
        
        # Execution Levels Summary
        ws['A25'] = "⏱️ Execution Levels"
        ws['A25'].font = Font(size=12, bold=True)
        
        # Analyze execution levels
        level_counts = Counter()
        for order in self.lg.transformation_orders.values():
            level_counts[order] += 1
        
        level_data = [["Execution Level", "Transformation Count", "Can Run in Parallel"]]
        for level in sorted(level_counts.keys()):
            count = level_counts[level]
            parallel = "Yes" if count > 1 else "No"
            level_data.append([f"Level {level}", count, parallel])
        
        for i, row in enumerate(level_data, 26):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 26:  # Header
                    cell.fill = self.colors['header']
                    cell.font = self.font_bold
                elif row[2] == "Yes":  # Parallel execution possible
                    cell.fill = self.colors['parallel']
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    def _create_detailed_lineage_sheet(self):
        """Create detailed lineage sheet with all connections"""
        ws = self.workbook.create_sheet("🔗 Detailed Lineage")
        
        # Prepare data
        lineage_data = []
        headers = [
            "Mapping Name", "From Field", "From Instance", "From Type", 
            "To Field", "To Instance", "To Type", "Execution Order", 
            "Execution Level", "Parallel Group", "Transformation Logic"
        ]
        lineage_data.append(headers)
        
        # Add connector data
        for connector in self.lg.connectors_data:
            row = [
                connector.mapping_name,
                connector.from_field,
                connector.from_instance,
                connector.from_instance_type,
                connector.to_field,
                connector.to_instance,
                connector.to_instance_type,
                connector.transformation_order,
                connector.execution_level,
                connector.parallel_group,
                connector.transformation_logic[:100] + "..." if len(connector.transformation_logic) > 100 else connector.transformation_logic
            ]
            lineage_data.append(row)
        
        # Write data to sheet
        for i, row in enumerate(lineage_data, 1):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                
                if i == 1:  # Header
                    cell.fill = self.colors['header']
                    cell.font = self.font_bold
                else:
                    # Color code by transformation type
                    if 'Source' in str(row[3]):
                        cell.fill = self.colors['source']
                    elif 'Target' in str(row[6]):
                        cell.fill = self.colors['target']
                    elif i % 2 == 0:
                        cell.fill = self.colors['light_gray']
        
        # Add filters
        ws.auto_filter.ref = f"A1:{chr(ord('A') + len(headers) - 1)}{len(lineage_data)}"
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    def _create_transformation_logic_sheet(self):
        """Create transformation logic sheet"""
        ws = self.workbook.create_sheet("⚙️ Transformation Logic")
        
        # Headers
        headers = ["Transformation Name", "Type", "Execution Order", "Business Logic", "Field Count"]
        for j, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=header)
            cell.fill = self.colors['header']
            cell.font = self.font_bold
        
        # Prepare transformation logic data
        logic_data = []
        
        if self.lg.mapping_name in self.lg.transformation_logic_cache:
            for trans_name, trans_info in self.lg.transformation_logic_cache[self.lg.mapping_name].items():
                # Get execution order
                full_trans_name = None
                for instance in self.lg.instances:
                    if instance.get('@NAME') == trans_name:
                        trans_type = instance.get('@TRANSFORMATION_TYPE', 'Unknown')
                        full_trans_name = f"{self.lg.create_transform_type_acronym(trans_type)}{trans_name}"
                        break
                
                exec_order = self.lg.transformation_orders.get(full_trans_name, 0)
                
                # Count fields in transformation
                field_count = 0
                for transformation in self.lg.transformations:
                    if transformation.get('@NAME') == trans_name:
                        transform_fields = transformation.get('TRANSFORMFIELD', [])
                        if isinstance(transform_fields, dict):
                            field_count = 1
                        elif isinstance(transform_fields, list):
                            field_count = len(transform_fields)
                        break
                
                logic_data.append([
                    trans_name,
                    trans_info['type'],
                    exec_order,
                    trans_info['logic'],
                    field_count
                ])
        
        # Sort by execution order
        logic_data.sort(key=lambda x: x[2])
        
        # Write data
        for i, row in enumerate(logic_data, 2):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                
                # Color code by transformation type
                if 'Expression' in str(row[1]):
                    cell.fill = self.colors['transformation']
                elif 'Source' in str(row[1]):
                    cell.fill = self.colors['source']
                elif 'Target' in str(row[1]):
                    cell.fill = self.colors['target']
                elif i % 2 == 0:
                    cell.fill = self.colors['light_gray']
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 80)
    
    def _create_execution_plan_sheet(self):
        """Create execution plan sheet with parallel groups and dependencies"""
        ws = self.workbook.create_sheet("📈 Execution Plan")
        
        # Title
        ws['A1'] = "Execution Plan & Parallel Groups"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Group transformations by execution level
        level_groups = defaultdict(list)
        for trans_name, order in self.lg.transformation_orders.items():
            parallel_group = self.lg.parallel_groups.get(trans_name, 'unknown')
            level_groups[order].append((trans_name, parallel_group))
        
        # Create execution plan table
        headers = ["Level", "Transformation", "Type", "Parallel Group", "Can Run Parallel", "Dependencies"]
        for j, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=j, value=header)
            cell.fill = self.colors['header']
            cell.font = self.font_bold
        
        row_num = 4
        G = self.lg.mapping_lineage_graph
        
        for level in sorted(level_groups.keys()):
            transformations = level_groups[level]
            
            # Group by parallel group within level
            pg_groups = defaultdict(list)
            for trans_name, pg in transformations:
                pg_groups[pg].append(trans_name)
            
            for pg, trans_list in pg_groups.items():
                for i, trans_name in enumerate(trans_list):
                    # Extract transformation type
                    trans_type = "Unknown"
                    clean_name = trans_name
                    for prefix in ['SRC_', 'TGT_', 'EXPR_', 'FILT_', 'AGGR_', 'SRCQ_', 'LOOK_', 'UPD_', 'CTR_']:
                        if trans_name.startswith(prefix):
                            trans_type = prefix.rstrip('_')
                            clean_name = trans_name[len(prefix):]
                            break
                    
                    # Get dependencies
                    dependencies = list(G.predecessors(trans_name)) if trans_name in G else []
                    dep_str = ", ".join([dep.split('_', 1)[-1] if '_' in dep else dep for dep in dependencies[:3]])
                    if len(dependencies) > 3:
                        dep_str += f" (+{len(dependencies)-3} more)"
                    
                    # Can run parallel
                    can_parallel = "Yes" if len(trans_list) > 1 else "No"
                    
                    # Write row
                    row_data = [level, clean_name, trans_type, pg.split('_')[-1], can_parallel, dep_str]
                    for j, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=j, value=value)
                        
                        # Color coding
                        if trans_type == "SRC":
                            cell.fill = self.colors['source']
                        elif trans_type == "TGT":
                            cell.fill = self.colors['target']
                        elif can_parallel == "Yes":
                            cell.fill = self.colors['parallel']
                        elif row_num % 2 == 0:
                            cell.fill = self.colors['light_gray']
                    
                    row_num += 1
        
        # Critical Path Analysis
        ws[f'A{row_num + 2}'] = "🎯 Critical Path Analysis"
        ws[f'A{row_num + 2}'].font = Font(size=12, bold=True)
        
        # Find longest path (critical path)
        try:
            if len(G.nodes()) > 0:
                # Find all paths from sources to sinks
                sources = [n for n in G.nodes() if G.in_degree(n) == 0]
                sinks = [n for n in G.nodes() if G.out_degree(n) == 0]
                
                longest_path = []
                max_length = 0
                
                for source in sources:
                    for sink in sinks:
                        try:
                            path = nx.shortest_path(G, source, sink)
                            if len(path) > max_length:
                                max_length = len(path)
                                longest_path = path
                        except nx.NetworkXNoPath:
                            continue
                
                if longest_path:
                    ws[f'A{row_num + 4}'] = f"Critical Path Length: {len(longest_path)} transformations"
                    ws[f'A{row_num + 5}'] = "Critical Path:"
                    
                    for i, node in enumerate(longest_path):
                        clean_node = node.split('_', 1)[-1] if '_' in node else node
                        ws[f'A{row_num + 6 + i}'] = f"  {i+1}. {clean_node}"
                        ws[f'A{row_num + 6 + i}'].fill = self.colors['critical']
        except Exception as e:
            ws[f'A{row_num + 4}'] = f"Critical path analysis error: {str(e)}"
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)
    
    def _create_data_flow_sheet(self):
        """Create data flow visualization sheet"""
        ws = self.workbook.create_sheet("🌊 Data Flow")
        
        ws['A1'] = "Data Flow Visualization"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Create ASCII-style flow diagram
        G = self.lg.mapping_lineage_graph
        
        # Group by execution levels
        level_groups = defaultdict(list)
        for trans_name, order in self.lg.transformation_orders.items():
            level_groups[order].append(trans_name)
        
        row_num = 3
        ws[f'A{row_num}'] = "Execution Flow (Top to Bottom):"
        ws[f'A{row_num}'].font = Font(size=12, bold=True)
        row_num += 2
        
        for level in sorted(level_groups.keys()):
            if level == 0:
                continue
                
            # Level header
            ws[f'A{row_num}'] = f"--- LEVEL {level} ---"
            ws[f'A{row_num}'].font = Font(bold=True)
            ws[f'A{row_num}'].fill = self.colors['header']
            row_num += 1
            
            # Show transformations at this level
            transformations = level_groups[level]
            
            if len(transformations) == 1:
                trans_name = transformations[0]
                clean_name = trans_name.split('_', 1)[-1] if '_' in trans_name else trans_name
                ws[f'A{row_num}'] = f"   [{clean_name}]"
                
                # Show connections
                predecessors = list(G.predecessors(trans_name))
                if predecessors:
                    pred_names = [p.split('_', 1)[-1] if '_' in p else p for p in predecessors]
                    ws[f'B{row_num}'] = f"← from: {', '.join(pred_names)}"
            else:
                ws[f'A{row_num}'] = f"   PARALLEL EXECUTION ({len(transformations)} transformations):"
                row_num += 1
                
                for trans_name in transformations:
                    clean_name = trans_name.split('_', 1)[-1] if '_' in trans_name else trans_name
                    ws[f'A{row_num}'] = f"     • [{clean_name}]"
                    
                    # Show connections
                    predecessors = list(G.predecessors(trans_name))
                    if predecessors:
                        pred_names = [p.split('_', 1)[-1] if '_' in p else p for p in predecessors]
                        ws[f'B{row_num}'] = f"← from: {', '.join(pred_names[:2])}"
                        if len(pred_names) > 2:
                            ws[f'B{row_num}'].value += f" (+{len(pred_names)-2})"
                    
                    # Color parallel transformations
                    ws[f'A{row_num}'].fill = self.colors['parallel']
                    row_num += 1
                
                row_num -= 1  # Adjust for the extra increment
            
            row_num += 2
        
        # Transformation Type Legend
        ws[f'A{row_num + 2}'] = "Legend:"
        ws[f'A{row_num + 2}'].font = Font(size=12, bold=True)
        
        legend_items = [
            ("Source Definition", "SRC_", self.colors['source']),
            ("Target Definition", "TGT_", self.colors['target']),
            ("Expression", "EXPR_", self.colors['transformation']),
            ("Filter", "FILT_", self.colors['transformation']),
            ("Aggregator", "AGGR_", self.colors['transformation']),
            ("Parallel Execution", "", self.colors['parallel'])
        ]
        
        for i, (desc, prefix, color) in enumerate(legend_items):
            row = row_num + 4 + i
            ws[f'A{row}'] = desc
            ws[f'B{row}'] = prefix
            ws[f'A{row}'].fill = color
            ws[f'B{row}'].fill = color
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 80)
    
    def _create_performance_analysis_sheet(self):
        """Create performance analysis sheet"""
        ws = self.workbook.create_sheet("📉 Performance Analysis")
        
        ws['A1'] = "Performance Analysis & Optimization Opportunities"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Analyze bottlenecks
        ws['A3'] = "🚨 Potential Bottlenecks"
        ws['A3'].font = Font(size=12, bold=True)
        
        G = self.lg.mapping_lineage_graph
        bottlenecks = []
        
        # Find nodes with high fan-in (convergence points)
        for node in G.nodes():
            in_degree = G.in_degree(node)
            out_degree = G.out_degree(node)
            
            if in_degree > 2:  # Convergence point
                clean_name = node.split('_', 1)[-1] if '_' in node else node
                bottlenecks.append(f"High fan-in: {clean_name} ({in_degree} inputs)")
            
            if out_degree > 3:  # High fan-out
                clean_name = node.split('_', 1)[-1] if '_' in node else node
                bottlenecks.append(f"High fan-out: {clean_name} ({out_degree} outputs)")
        
        for i, bottleneck in enumerate(bottlenecks, 5):
            ws[f'A{i}'] = bottleneck
            ws[f'A{i}'].fill = self.colors['critical']
        
        # Parallel execution opportunities
        row_start = 5 + len(bottlenecks) + 2
        ws[f'A{row_start}'] = "⚡ Parallel Execution Opportunities"
        ws[f'A{row_start}'].font = Font(size=12, bold=True)
        
        # Analyze parallel groups
        parallel_stats = defaultdict(int)
        for pg in self.lg.parallel_groups.values():
            if 'group_' in pg:
                parallel_stats[pg] += 1
        
        parallel_opportunities = [(pg, count) for pg, count in parallel_stats.items() if count > 1]
        parallel_opportunities.sort(key=lambda x: x[1], reverse=True)
        
        for i, (pg, count) in enumerate(parallel_opportunities):
            row = row_start + 2 + i
            ws[f'A{row}'] = f"Parallel Group {pg.split('_')[-1]}: {count} transformations can run simultaneously"
            ws[f'A{row}'].fill = self.colors['parallel']
        
        # Resource utilization analysis
        row_start = row_start + len(parallel_opportunities) + 4
        ws[f'A{row_start}'] = "💾 Resource Utilization Analysis"
        ws[f'A{row_start}'].font = Font(size=12, bold=True)
        
        # Count transformation types
        type_counts = defaultdict(int)
        for instance in self.lg.instances:
            trans_type = instance.get('@TRANSFORMATION_TYPE', 'Unknown')
            type_counts[trans_type] += 1
        
        # Transformation type analysis
        analysis_data = [["Transformation Type", "Count", "Resource Impact", "Optimization Notes"]]
        
        impact_map = {
            'Source Definition': ('Low', 'Consider partitioning for large tables'),
            'Target Definition': ('Medium', 'Optimize commit intervals and bulk loading'),
            'Expression': ('Low', 'Combine multiple expressions where possible'),
            'Filter': ('Low', 'Push filters close to sources'),
            'Aggregator': ('High', 'Consider pre-aggregating or partitioning'),
            'Lookup Procedure': ('High', 'Cache frequently used lookups'),
            'Source Qualifier': ('Medium', 'Optimize SQL and use pushdown optimization'),
            'Custom Transformation': ('Variable', 'Review custom logic for efficiency')
        }
        
        for trans_type, count in type_counts.items():
            impact, notes = impact_map.get(trans_type, ('Unknown', 'Review transformation logic'))
            analysis_data.append([trans_type, count, impact, notes])
        
        # Write analysis table
        for i, row in enumerate(analysis_data, row_start + 2):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                
                if i == row_start + 2:  # Header
                    cell.fill = self.colors['header']
                    cell.font = self.font_bold
                elif len(row) > 2 and row[2] == 'High':  # High impact
                    cell.fill = self.colors['critical']
                elif len(row) > 2 and row[2] == 'Medium':  # Medium impact
                    cell.fill = self.colors['transformation']
                elif i % 2 == 0:
                    cell.fill = self.colors['light_gray']
        
        # Critical path impact
        row_start = row_start + len(analysis_data) + 4
        ws[f'A{row_start}'] = "⏰ Critical Path Impact"
        ws[f'A{row_start}'].font = Font(size=12, bold=True)
        
        max_level = max(self.lg.transformation_orders.values()) if self.lg.transformation_orders else 0
        ws[f'A{row_start + 2}'] = f"Total Execution Levels: {max_level}"
        ws[f'A{row_start + 3}'] = f"Estimated Minimum Execution Time: {max_level} sequential steps"
        
        # Calculate parallel efficiency
        total_transformations = len([t for t in self.lg.transformation_orders.values() if t > 0])
        if max_level > 0:
            parallel_efficiency = (total_transformations / max_level) * 100
            ws[f'A{row_start + 4}'] = f"Parallel Efficiency: {parallel_efficiency:.1f}% (higher is better)"
            
            if parallel_efficiency > 75:
                ws[f'A{row_start + 5}'] = "✅ Good parallelization - efficient mapping design"
                ws[f'A{row_start + 5}'].fill = self.colors['source']
            elif parallel_efficiency > 50:
                ws[f'A{row_start + 5}'] = "⚠️ Moderate parallelization - some optimization possible"
                ws[f'A{row_start + 5}'].fill = self.colors['transformation']
            else:
                ws[f'A{row_start + 5}'] = "🔴 Low parallelization - significant optimization needed"
                ws[f'A{row_start + 5}'].fill = self.colors['critical']
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


# Utility function to export mapping to Excel
def export_mapping_to_excel(lineage_generator, output_path: str):
    """
    Export InfaLineageGenerator data to comprehensive Excel report
    
    Args:
        lineage_generator: InfaLineageGenerator instance
        output_path: Path for output Excel file
    """
    exporter = InfaLineageExcelExporter(lineage_generator)
    exporter.generate_excel_report(output_path)
    
    print(f"\n📊 Excel Report Generated Successfully!")
    print(f"📁 Location: {output_path}")
    print(f"📋 Contains 6 comprehensive sheets:")
    print(f"   • Overview: Mapping statistics and component analysis")
    print(f"   • Detailed Lineage: Field-level connection details")
    print(f"   • Transformation Logic: Business rules and logic")
    print(f"   • Execution Plan: Parallel groups and dependencies")
    print(f"   • Data Flow: Visual flow representation")
    print(f"   • Performance Analysis: Bottlenecks and optimization tips") 